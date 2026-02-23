#!/usr/bin/env python3
"""
Script to create the GreenStyle Financial Model Excel file
This creates a comprehensive financial model with German headings
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_finanzmodell():
    """Create the complete financial model Excel file"""
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create all sheets
    create_annahmen_sheet(wb)
    create_umsatzplanung_sheet(wb)
    create_kostenplanung_sheet(wb)
    create_guv_sheet(wb)
    create_liquiditaet_sheet(wb)
    create_bilanz_sheet(wb)
    create_investitionsplan_sheet(wb)
    create_kapitalbedarf_sheet(wb)
    create_breakeven_sheet(wb)
    create_sensitivitaet_sheet(wb)
    create_schuldendienst_sheet(wb)
    create_dashboard_sheet(wb)
    
    # Save file
    filename = 'finanzmodell-greenstyle.xlsx'
    wb.save(filename)
    print(f"✓ Finanzmodell erstellt: {filename}")
    return filename

def create_annahmen_sheet(wb):
    """Create assumptions sheet"""
    ws = wb.create_sheet("Annahmen", 0)
    
    # Header
    ws['A1'] = "FINANZMODELL GREENSTYLE"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = "Planungsannahmen"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A3'] = f"Stand: {datetime.now().strftime('%d.%m.%Y')}"
    
    # Umsatzannahmen
    row = 5
    ws[f'A{row}'] = "UMSATZANNAHMEN"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    assumptions = [
        ("", "Jahr 1", "Jahr 2", "Jahr 3"),
        ("Anzahl Kunden", 2000, 3800, 5700),
        ("Durchschn. Warenkorbwert (€)", 142, 137, 137),
        ("Käufe pro Kunde", 1.0, 1.4, 1.4),
        ("Wiederkaufrate", "0%", "30%", "40%"),
        ("Monatl. Wachstum (Ø)", "12%", "8%", "6%"),
        ("", "", "", ""),
        ("Berechneter Umsatz (€)", 285000, 520000, 780000),
    ]
    
    for i, row_data in enumerate(assumptions, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1:
                cell.font = Font(bold=True)
    
    # Kostenstruktur
    row = row + len(assumptions) + 2
    ws[f'A{row}'] = "KOSTENSTRUKTUR"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    costs = [
        ("", "% vom Umsatz", "Fixbetrag/Jahr"),
        ("Wareneinsatz (COGS)", "50%", ""),
        ("Versandkosten", "2,5%", ""),
        ("Retourenkosten", "4,0%", ""),
        ("Zahlungsgebühren", "2,5%", ""),
        ("Verpackung", "1,0%", ""),
        ("", "", ""),
        ("Marketing", "8,8% / 10,7% / 11,6%", ""),
        ("Software/IT", "", "2.160 €"),
        ("Steuerberatung", "", "3.000 €"),
        ("Versicherungen", "", "1.400 €"),
    ]
    
    for i, row_data in enumerate(costs, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1:
                cell.font = Font(bold=True)
    
    # Finanzierung
    row = row + len(costs) + 2
    ws[f'A{row}'] = "FINANZIERUNG"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    financing = [
        ("", "Betrag (€)", "Zinssatz", "Laufzeit"),
        ("Eigenkapital", 25000, "-", "-"),
        ("KfW-Kredit", 80000, "4,5%", "10 Jahre"),
        ("Sparkassen-Kredit", 20000, "6,0%", "5 Jahre"),
        ("GESAMT", 125000, "", ""),
    ]
    
    for i, row_data in enumerate(financing, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1 or i == row+len(financing):
                cell.font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def create_umsatzplanung_sheet(wb):
    """Create sales planning sheet"""
    ws = wb.create_sheet("Umsatzplanung")
    
    ws['A1'] = "UMSATZPLANUNG"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Jahr 1 - Monatlich
    ws['A3'] = "Jahr 1 (2025) - Monatlich"
    ws['A3'].font = Font(bold=True)
    
    headers = ["Monat", "Bestellungen", "Ø Warenkorb (€)", "Umsatz (€)", "Kumuliert (€)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    months_data = [
        ("Apr", 50, 150, 7500, 7500),
        ("Mai", 75, 145, 10875, 18375),
        ("Jun", 110, 143, 15730, 34105),
        ("Jul", 150, 142, 21300, 55405),
        ("Aug", 200, 140, 28000, 83405),
        ("Sep", 250, 140, 35000, 118405),
        ("Okt", 300, 140, 42000, 160405),
        ("Nov", 350, 140, 49000, 209405),
        ("Dez", 400, 142, 56800, 266205),
        ("Jan", 100, 138, 13800, 280005),
        ("Feb", 35, 143, 5005, 285010),
    ]
    
    for i, row_data in enumerate(months_data, start=5):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)
    
    # Total
    total_row = 5 + len(months_data)
    ws[f'A{total_row}'] = "GESAMT Jahr 1"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'B{total_row}'] = 2020
    ws[f'C{total_row}'] = 141
    ws[f'D{total_row}'] = 285010
    ws[f'E{total_row}'] = 285010
    
    # Jahr 2 & 3
    row = total_row + 3
    ws[f'A{row}'] = "Jahr 2-3 - Quartalsweise"
    ws[f'A{row}'].font = Font(bold=True)
    
    quarterly_headers = ["Periode", "Bestellungen", "Ø Warenkorb (€)", "Umsatz (€)"]
    for col, header in enumerate(quarterly_headers, start=1):
        cell = ws.cell(row=row+1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    quarterly_data = [
        ("2026 Q1", 800, 135, 108000),
        ("2026 Q2", 950, 137, 130150),
        ("2026 Q3", 1100, 138, 151800),
        ("2026 Q4", 1300, 138, 179400),
        ("Jahr 2 GESAMT", 4150, 137, 569350),
        ("", "", "", ""),
        ("2027 Q1", 1400, 137, 191800),
        ("2027 Q2", 1600, 137, 219200),
        ("2027 Q3", 1800, 137, 246600),
        ("2027 Q4", 2100, 137, 287700),
        ("Jahr 3 GESAMT", 6900, 137, 945300),
    ]
    
    for i, row_data in enumerate(quarterly_data, start=row+2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if "GESAMT" in str(row_data[0]):
                cell.font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

def create_kostenplanung_sheet(wb):
    """Create cost planning sheet"""
    ws = wb.create_sheet("Kostenplanung")
    
    ws['A1'] = "KOSTENPLANUNG (3 Jahre)"
    ws['A1'].font = Font(size=14, bold=True)
    
    headers = ["Kostenart", "Jahr 1 (€)", "Jahr 2 (€)", "Jahr 3 (€)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    costs = [
        ("VARIABLE KOSTEN", "", "", ""),
        ("Wareneinsatz (COGS)", 142500, 260000, 390000),
        ("Versandkosten", 7125, 13000, 19500),
        ("Retourenkosten", 11400, 20800, 31200),
        ("Zahlungsgebühren", 7125, 13000, 19500),
        ("Verpackung", 2850, 5200, 7800),
        ("Summe variabel", 171000, 312000, 468000),
        ("", "", "", ""),
        ("FIXKOSTEN", "", "", ""),
        ("Personalkosten", 0, 12600, 38400),
        ("Marketing", 25000, 55800, 90200),
        ("Miete/Büro", 0, 3600, 9600),
        ("Software/IT", 2160, 2400, 2880),
        ("Steuerberatung", 3000, 3600, 4200),
        ("Versicherungen", 1400, 2600, 3000),
        ("Rechtsberatung", 1500, 1000, 1000),
        ("Sonstiges", 2000, 3000, 4000),
        ("Summe fix", 35060, 84600, 153280),
        ("", "", "", ""),
        ("GESAMTKOSTEN", 206060, 396600, 621280),
    ]
    
    for i, row_data in enumerate(costs, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if row_data[0] in ["VARIABLE KOSTEN", "FIXKOSTEN", "Summe variabel", "Summe fix", "GESAMTKOSTEN"]:
                cell.font = Font(bold=True)
                if row_data[0] in ["VARIABLE KOSTEN", "FIXKOSTEN"]:
                    cell.fill = PatternFill(start_color="E7E6E6", fill_type="solid")
    
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 15

def create_guv_sheet(wb):
    """Create P&L sheet"""
    ws = wb.create_sheet("GuV")
    
    ws['A1'] = "GEWINN- UND VERLUSTRECHNUNG (GuV)"
    ws['A1'].font = Font(size=14, bold=True)
    
    headers = ["Position", "Jahr 1 (€)", "Jahr 2 (€)", "Jahr 3 (€)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    guv_data = [
        ("Umsatzerlöse", 285000, 520000, 780000),
        ("Versandeinnahmen", 3000, 5000, 7000),
        ("Gesamterlöse", 288000, 525000, 787000),
        ("", "", "", ""),
        ("Wareneinsatz", -142500, -260000, -390000),
        ("BRUTTOGEWINN", 145500, 265000, 397000),
        ("Bruttomarge (%)", "50,5%", "50,5%", "50,4%"),
        ("", "", "", ""),
        ("BETRIEBSAUSGABEN", "", "", ""),
        ("Versandkosten", -7125, -13000, -19500),
        ("Retourenkosten", -11400, -20800, -31200),
        ("Zahlungsgebühren", -7125, -13000, -19500),
        ("Verpackung", -2850, -5200, -7800),
        ("Personalkosten", 0, -12600, -38400),
        ("Marketing", -25000, -55800, -90200),
        ("Miete/Büro", 0, -3600, -9600),
        ("Software/IT", -2160, -2400, -2880),
        ("Steuerberatung", -3000, -3600, -4200),
        ("Versicherungen", -1400, -2600, -3000),
        ("Rechtsberatung", -1500, -1000, -1000),
        ("Sonstiges", -2000, -3000, -4000),
        ("Summe Betriebsausgaben", -63560, -136600, -231280),
        ("", "", "", ""),
        ("EBITDA", 81940, 128400, 165720),
        ("EBITDA-Marge (%)", "28,5%", "24,5%", "21,0%"),
        ("", "", "", ""),
        ("Abschreibungen", -7475, -7475, -7475),
        ("EBIT", 74465, 120925, 158245),
        ("", "", "", ""),
        ("Zinsaufwand", -5400, -5400, -5400),
        ("EBT (Gewinn vor Steuern)", 69065, 115525, 152845),
        ("", "", "", ""),
        ("Steuern", 0, 0, -45854),
        ("JAHRESÜBERSCHUSS", 69065, 115525, 106991),
        ("Nettomarge (%)", "24,0%", "22,0%", "13,6%"),
    ]
    
    for i, row_data in enumerate(guv_data, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if row_data[0] in ["Gesamterlöse", "BRUTTOGEWINN", "EBITDA", "EBIT", "EBT (Gewinn vor Steuern)", "JAHRESÜBERSCHUSS"]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")
            elif row_data[0] in ["BETRIEBSAUSGABEN"]:
                cell.font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 15

def create_liquiditaet_sheet(wb):
    """Create liquidity/cash flow sheet"""
    ws = wb.create_sheet("Liquidität")
    
    ws['A1'] = "LIQUIDITÄTSPLANUNG (Cash Flow)"
    ws['A1'].font = Font(size=14, bold=True)
    
    # 3-year overview
    headers = ["Position", "Jahr 1 (€)", "Jahr 2 (€)", "Jahr 3 (€)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    cashflow = [
        ("OPERATIVER CASH FLOW", "", "", ""),
        ("Jahresüberschuss", 69065, 115525, 106991),
        ("+ Abschreibungen", 7475, 7475, 7475),
        ("- Lagererweiterung", -10000, -15000, -20000),
        ("= Operativer CF", 66540, 108000, 94466),
        ("", "", "", ""),
        ("INVESTITIONS-CF", "", "", ""),
        ("Investitionen", -14950, -5000, -8000),
        ("= Investitions-CF", -14950, -5000, -8000),
        ("", "", "", ""),
        ("FINANZIERUNGS-CF", "", "", ""),
        ("Eigenkapital", 25000, 0, 0),
        ("Kredite", 100000, 0, 0),
        ("Tilgung", 0, -5000, -15000),
        ("Zinsen", -5400, -5400, -5400),
        ("= Finanzierungs-CF", 119600, -10400, -20400),
        ("", "", "", ""),
        ("GESAMT-CASH FLOW", 171190, 92600, 66066),
        ("Liquidität (kumuliert)", 171190, 263790, 329856),
    ]
    
    for i, row_data in enumerate(cashflow, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if row_data[0] in ["OPERATIVER CASH FLOW", "INVESTITIONS-CF", "FINANZIERUNGS-CF"]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", fill_type="solid")
            elif "=" in row_data[0] or row_data[0] in ["GESAMT-CASH FLOW", "Liquidität (kumuliert)"]:
                cell.font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 15

def create_bilanz_sheet(wb):
    """Create balance sheet"""
    ws = wb.create_sheet("Bilanz")
    
    ws['A1'] = "BILANZPLANUNG"
    ws['A1'].font = Font(size=14, bold=True)
    
    headers = ["Position", "31.12.2025", "31.12.2026", "31.12.2027"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    bilanz = [
        ("AKTIVA", "", "", ""),
        ("", "", "", ""),
        ("Anlagevermögen", "", "", ""),
        ("Immaterielle VG", 7475, 4983, 2492),
        ("Sachanlagen", 0, 3333, 5333),
        ("Summe AV", 7475, 8316, 7825),
        ("", "", "", ""),
        ("Umlaufvermögen", "", "", ""),
        ("Vorräte (Warenlager)", 55000, 70000, 90000),
        ("Forderungen", 5000, 9000, 13000),
        ("Liquide Mittel", 171190, 263790, 329856),
        ("Summe UV", 231190, 342790, 432856),
        ("", "", "", ""),
        ("SUMME AKTIVA", 238665, 351106, 440681),
        ("", "", "", ""),
        ("", "", "", ""),
        ("PASSIVA", "", "", ""),
        ("", "", "", ""),
        ("Eigenkapital", "", "", ""),
        ("Stammkapital", 25000, 25000, 25000),
        ("Gewinnrücklage", 0, 69065, 184590),
        ("Jahresüberschuss", 69065, 115525, 106991),
        ("Summe EK", 94065, 209590, 316581),
        ("", "", "", ""),
        ("Fremdkapital", "", "", ""),
        ("Langfristige Verbindlichkeiten", 100000, 95000, 80000),
        ("Kurzfristige Verbindlichkeiten", 44600, 46516, 44100),
        ("Summe FK", 144600, 141516, 124100),
        ("", "", "", ""),
        ("SUMME PASSIVA", 238665, 351106, 440681),
        ("", "", "", ""),
        ("KENNZAHLEN", "", "", ""),
        ("Eigenkapitalquote (%)", "39,4%", "59,7%", "71,8%"),
        ("Verschuldungsgrad (%)", "153,7%", "67,5%", "39,2%"),
    ]
    
    for i, row_data in enumerate(bilanz, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if row_data[0] in ["AKTIVA", "PASSIVA", "KENNZAHLEN"]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
            elif row_data[0] in ["SUMME AKTIVA", "SUMME PASSIVA", "Summe AV", "Summe UV", "Summe EK", "Summe FK"]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 15

def create_investitionsplan_sheet(wb):
    """Create investment plan sheet"""
    ws = wb.create_sheet("Investitionsplan")
    
    ws['A1'] = "INVESTITIONSPLAN"
    ws['A1'].font = Font(size=14, bold=True)
    
    ws['A3'] = "Gründungsinvestitionen"
    ws['A3'].font = Font(bold=True)
    
    investments = [
        ("Position", "Betrag (€)", "Abschreibung"),
        ("Rechtsberatung, Notar, HR", 2450, "Sofort"),
        ("Website-Entwicklung", 8000, "3 Jahre"),
        ("Branding (Logo, Design)", 3000, "3 Jahre"),
        ("Erstausstattung Büro", 1500, "3 Jahre"),
        ("GESAMT", 14950, ""),
    ]
    
    for i, row_data in enumerate(investments, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 4 or i == 4 + len(investments) - 1:
                cell.font = Font(bold=True)
    
    ws['A11'] = "Warenlager (Erstausstattung)"
    ws['A11'].font = Font(bold=True)
    ws['A12'] = "Warenlager"
    ws['B12'] = 45000
    ws['C12'] = "Umlaufvermögen"
    
    ws['A14'] = "Betriebsmittel"
    ws['A14'].font = Font(bold=True)
    ws['A15'] = "Liquiditätsreserve (6 Monate)"
    ws['B15'] = 30000
    
    ws['A17'] = "GESAMTINVESTITION"
    ws['A17'].font = Font(bold=True, size=11)
    ws['B17'] = 89950
    ws['B17'].font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20

def create_kapitalbedarf_sheet(wb):
    """Create capital requirements sheet"""
    ws = wb.create_sheet("Kapitalbedarf")
    
    ws['A1'] = "KAPITALBEDARFSPLAN"
    ws['A1'].font = Font(size=14, bold=True)
    
    headers = ["Verwendung", "Betrag (€)", "Anteil (%)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    capital = [
        ("Gründungskosten", 14950, "12%"),
        ("Warenlager (Erstausstattung)", 45000, "36%"),
        ("Liquiditätsreserve (6 Monate)", 30000, "24%"),
        ("Marketing (Launch)", 10000, "8%"),
        ("Laufende Kosten (3 Monate)", 15000, "12%"),
        ("Puffer", 10050, "8%"),
        ("GESAMTKAPITALBEDARF", 125000, "100%"),
    ]
    
    for i, row_data in enumerate(capital, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 4 + len(capital) - 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    
    row = 4 + len(capital) + 2
    ws[f'A{row}'] = "FINANZIERUNG"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    financing = [
        ("Quelle", "Betrag (€)", "Anteil (%)"),
        ("Eigenkapital (Gründerin)", 25000, "20%"),
        ("KfW-Gründerkredit", 80000, "64%"),
        ("Sparkassen-Betriebsmittelkredit", 20000, "16%"),
        ("GESAMTFINANZIERUNG", 125000, "100%"),
    ]
    
    for i, row_data in enumerate(financing, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1 or i == row + len(financing):
                cell.font = Font(bold=True)
                if i == row + len(financing):
                    cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def create_breakeven_sheet(wb):
    """Create break-even analysis sheet"""
    ws = wb.create_sheet("Break-Even")
    
    ws['A1'] = "BREAK-EVEN-ANALYSE"
    ws['A1'].font = Font(size=14, bold=True)
    
    ws['A3'] = "Deckungsbeitragsrechnung"
    ws['A3'].font = Font(bold=True)
    
    data = [
        ("Position", "Wert"),
        ("Durchschn. Warenkorbwert", "142 €"),
        ("Variable Kosten pro Verkauf", "85 €"),
        ("Deckungsbeitrag pro Verkauf", "57 €"),
        ("Deckungsbeitrag in %", "40%"),
        ("", ""),
        ("Fixkosten pro Monat (Ø)", "2.922 €"),
        ("", ""),
        ("Break-Even-Umsatz/Monat", "7.305 €"),
        ("Break-Even-Bestellungen/Monat", "51 Stück"),
        ("", ""),
        ("Erreicht in", "Monat 1 (April 2025)"),
    ]
    
    for i, row_data in enumerate(data, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 4 or "Break-Even" in str(row_data[0]):
                cell.font = Font(bold=True)
    
    ws['A18'] = "Kumulierter Break-Even (inkl. Investitionen)"
    ws['A18'].font = Font(bold=True)
    ws['A19'] = "Gesamtinvestitionen"
    ws['B19'] = "125.000 €"
    ws['A20'] = "Break-Even erreicht"
    ws['B20'] = "Monat 18 (September 2026)"
    ws['B20'].font = Font(bold=True)
    ws['B20'].fill = PatternFill(start_color="C6E0B4", fill_type="solid")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

def create_sensitivitaet_sheet(wb):
    """Create sensitivity analysis sheet"""
    ws = wb.create_sheet("Sensitivität")
    
    ws['A1'] = "SENSITIVITÄTSANALYSE"
    ws['A1'].font = Font(size=14, bold=True)
    
    ws['A3'] = "Szenarien (Jahr 1)"
    ws['A3'].font = Font(bold=True)
    
    scenarios = [
        ("Szenario", "Umsatz (€)", "Jahresüberschuss (€)", "Abweichung"),
        ("Worst Case (-30%)", 199500, 34565, "-50%"),
        ("Pessimistisch (-20%)", 228000, 46265, "-33%"),
        ("Konservativ (-10%)", 256500, 57665, "-17%"),
        ("Base Case (Plan)", 285000, 69065, "0%"),
        ("Optimistisch (+10%)", 313500, 80465, "+16%"),
        ("Best Case (+20%)", 342000, 91865, "+33%"),
        ("Sehr optimistisch (+30%)", 370500, 103265, "+50%"),
    ]
    
    for i, row_data in enumerate(scenarios, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 4:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
            elif "Base Case" in str(row_data[0]):
                cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    
    row = 4 + len(scenarios) + 2
    ws[f'A{row}'] = "Sensitivität Wareneinsatz (COGS)"
    ws[f'A{row}'].font = Font(bold=True)
    
    cogs_sens = [
        ("COGS-Abweichung", "Bruttogewinn (€)", "Jahresüberschuss (€)"),
        ("+10%", 131250, 54815),
        ("+5%", 138375, 61940),
        ("0% (Base)", 145500, 69065),
        ("-5%", 152625, 76190),
        ("-10%", 159750, 83315),
    ]
    
    for i, row_data in enumerate(cogs_sens, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
            elif "0%" in str(row_data[0]):
                cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 20

def create_schuldendienst_sheet(wb):
    """Create debt service sheet"""
    ws = wb.create_sheet("Schuldendienst")
    
    ws['A1'] = "SCHULDENDIENSTFÄHIGKEIT"
    ws['A1'].font = Font(size=14, bold=True)
    
    ws['A3'] = "Debt Service Coverage Ratio (DSCR)"
    ws['A3'].font = Font(bold=True)
    
    dscr = [
        ("Jahr", "EBITDA (€)", "Steuern (€)", "Tilgung (€)", "Zinsen (€)", "Schuldendienst (€)", "DSCR"),
        (2025, 81940, 0, 0, 4800, 4800, 17.1),
        (2026, 128400, 0, 5000, 4800, 9800, 13.1),
        (2027, 165720, 45854, 15000, 4500, 19500, 6.1),
    ]
    
    for i, row_data in enumerate(dscr, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 4:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    ws['A10'] = "Bewertung: DSCR > 1,2 = Sehr gut (Bank-Mindestanforderung)"
    ws['A10'].font = Font(italic=True)
    ws['A11'] = "Alle Jahre deutlich über Mindestanforderung ✓"
    ws['A11'].font = Font(bold=True)
    ws['A11'].fill = PatternFill(start_color="C6E0B4", fill_type="solid")
    
    row = 14
    ws[f'A{row}'] = "Tilgungsplan KfW-Kredit (80.000 €)"
    ws[f'A{row}'].font = Font(bold=True)
    
    kfw = [
        ("Jahr", "Restschuld Anfang", "Tilgung", "Zinsen", "Rate", "Restschuld Ende"),
        (2025, 80000, 0, 3600, 3600, 80000),
        (2026, 80000, 0, 3600, 3600, 80000),
        (2027, 80000, 10000, 3600, 13600, 70000),
        (2028, 70000, 10000, 3150, 13150, 60000),
        (2029, 60000, 10000, 2700, 12700, 50000),
    ]
    
    for i, row_data in enumerate(kfw, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    row = row + len(kfw) + 2
    ws[f'A{row}'] = "Tilgungsplan Sparkassen-Kredit (20.000 €)"
    ws[f'A{row}'].font = Font(bold=True)
    
    sparkasse = [
        ("Jahr", "Restschuld Anfang", "Tilgung", "Zinsen", "Rate", "Restschuld Ende"),
        (2025, 20000, 0, 1200, 1200, 20000),
        (2026, 20000, 5000, 1200, 6200, 15000),
        (2027, 15000, 5000, 900, 5900, 10000),
        (2028, 10000, 5000, 600, 5600, 5000),
        (2029, 5000, 5000, 300, 5300, 0),
    ]
    
    for i, row_data in enumerate(sparkasse, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18

def create_dashboard_sheet(wb):
    """Create dashboard/summary sheet"""
    ws = wb.create_sheet("Dashboard", 0)
    
    ws['A1'] = "GREENSTYLE - FINANZMODELL DASHBOARD"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = "Nachhaltige Mode Online - Businessplan 2025-2027"
    ws['A2'].font = Font(size=11)
    
    ws['A4'] = "KEY PERFORMANCE INDICATORS"
    ws['A4'].font = Font(size=12, bold=True)
    ws['A4'].fill = PatternFill(start_color="4472C4", fill_type="solid")
    ws['A4'].font = Font(size=12, bold=True, color="FFFFFF")
    
    kpis = [
        ("Kennzahl", "Jahr 1", "Jahr 2", "Jahr 3"),
        ("Umsatz (€)", "285.000", "520.000", "780.000"),
        ("EBITDA (€)", "81.940", "128.400", "165.720"),
        ("Jahresüberschuss (€)", "69.065", "115.525", "106.991"),
        ("EBITDA-Marge (%)", "28,5%", "24,5%", "21,0%"),
        ("Nettomarge (%)", "24,0%", "22,0%", "13,6%"),
        ("", "", "", ""),
        ("Kunden", "2.000", "3.800", "5.700"),
        ("Ø Warenkorbwert (€)", "142", "137", "137"),
        ("Bestellungen", "2.020", "4.150", "6.900"),
        ("", "", "", ""),
        ("Liquidität (€)", "171.190", "263.790", "329.856"),
        ("Eigenkapitalquote (%)", "39,4%", "59,7%", "71,8%"),
        ("DSCR", "17,1", "13,1", "6,1"),
    ]
    
    for i, row_data in enumerate(kpis, start=5):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 5:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
            elif row_data[0] in ["Umsatz (€)", "EBITDA (€)", "Jahresüberschuss (€)"]:
                cell.font = Font(bold=True)
    
    row = 5 + len(kpis) + 2
    ws[f'A{row}'] = "FINANZIERUNG"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", fill_type="solid")
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    
    financing = [
        ("Quelle", "Betrag (€)", "Anteil"),
        ("Eigenkapital", "25.000", "20%"),
        ("KfW-Kredit", "80.000", "64%"),
        ("Sparkassen-Kredit", "20.000", "16%"),
        ("GESAMT", "125.000", "100%"),
    ]
    
    for i, row_data in enumerate(financing, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1 or i == row+len(financing):
                cell.font = Font(bold=True)
    
    row = row + len(financing) + 2
    ws[f'A{row}'] = "MEILENSTEINE"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", fill_type="solid")
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    
    milestones = [
        ("Meilenstein", "Zeitpunkt", "Status"),
        ("Operativer Break-Even", "Monat 1 (Apr 2025)", "✓"),
        ("Kumulierter Break-Even", "Monat 18 (Sep 2026)", "Geplant"),
        ("Profitabilität", "Jahr 1", "Geplant"),
        ("Marktplatz-Integration", "Jahr 2", "Geplant"),
    ]
    
    for i, row_data in enumerate(milestones, start=row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == row+1:
                cell.font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 15

if __name__ == "__main__":
    print("Erstelle GreenStyle Finanzmodell...")
    filename = create_finanzmodell()
    print(f"\n✓ Fertig! Datei erstellt: {filename}")
    print("\nDas Finanzmodell enthält folgende Tabellenblätter:")
    print("  1. Dashboard (Übersicht)")
    print("  2. Annahmen")
    print("  3. Umsatzplanung")
    print("  4. Kostenplanung")
    print("  5. GuV (Gewinn- und Verlustrechnung)")
    print("  6. Liquidität (Cash Flow)")
    print("  7. Bilanz")
    print("  8. Investitionsplan")
    print("  9. Kapitalbedarf")
    print(" 10. Break-Even")
    print(" 11. Sensitivität")
    print(" 12. Schuldendienst")
